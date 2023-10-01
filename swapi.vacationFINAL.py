#-----------------NOTES-----------------
# use customertkinter module for a more sleek GUI
# Could try pyqt. Qt creator
# could try streamlit if you are building a data app
# someone recommended kivy




#-----------------IMPORTS-----------------

import requests, json
from openpyxl import Workbook

import tkinter as tk
from tkinter import ttk
import tkinter.ttk as ttk
from tkinter import simpledialog #for Treeview Widget (excel spreadsheet appearance in notebook tabs)
#-----------------API-----------------

planetAPI = 'https://swapi.dev/api/planets/' #60 Planets
starshipAPI = 'https://swapi.dev/api/starships/' #36 Starships
vehicleAPI = 'https://swapi.dev/api/vehicles/' #39 Vehicles

#SEPARATOR: used to group related widgets in a window. It displays a horizontal or vertical line that contains a label or a small amount of text describing the purpose of the group of widgets. NOTE: Instructions on top, or blurb about selection and then the selection options below.
    #CODE: Place this between the instructions and the selection options within each notebook tab.
    #separator = ttk.Separator(frameNAME, orient = tk.HORIZONTAL)
    #separator.pack(expand = True, fill = tk.X)


#BUTTON (PUSHBUTTON): executes a command or displays a message when the user clicks it

#RADIOBUTTON (RadioButton): allows only one option to be selected by the user

#ENTRY (TEXTBOX): allows the user to enter a single line of text

#FRAME (CONTAINER): a rectangular region used to group related widgets or provide padding between widgets

#LISTBOX: displays a list of options to a user

#MENUBUTTON: displays a menu in the application when clicked

#TEXT: displays multiple lines of text to the user

#PANNEDWINDOW (CONTAINER): a container that contains two sub-containers arranged horizontally or vertically

#TKMESSAGEBOX (MESSAGEBOX): displays a message box that can contain text, buttons, and icons

#SIMPLEDIALOG MODULE
#askinteger(title, prompt, **kw) --accepts an integer input from the user.p
#add to imports:
#from tkinter.simpledialog import askinteger
# from tkinter import *
# from tkinter import messagebox
# top = Tk()
#CODE
#top.geometry("100x100")
#def show():
#    num = askinteger("input", "Input an Integer")
#    print(num)
#B = Button(top, text = "Next", command = show)
#B.place(x=50, y=50)
#top.mainloop()

#ttk MODULE
#add to imports:
#from tkinter import *
#from tkinter.ttk import * 

#imports: button, checkbutton, entry, frame, label, labelfram, menubutton, panedwindow, radiobutton, scale, and scrollbar. Use the ttk.style class to create and manage your own widget styles(improved style effects).

#TREEVIEW widget: displays a hierarchical collection of items using columns. Each item has a textual label, an optional image, and an optional list of data values. The data values are displayed in successive columns after the tree label. (MIGHT BE OUR BEST BET FOR THE GUI) PLACE WITHIN THE NOTEBOOK WIDGET. SO: NOTEBOOK > FRAME > TREEVIEW. Or Each notebook tab has a treeview widget that displays the data for that tab.
# #-----------------OPENPYXL HEADER SETUP-----------------



# #Starship Worksheet
# headers2 = ['Starship', 'Type of Ship', 'Price', 'Maximum Speed', 'Total Seats', 'Crew Size', 'Ship Size', 'Passenger Experience']
# for index, header in enumerate(headers2, 1):
#     starship_ws.cell(row=1, column=index, value=header)

# #Vehicle Worksheet
# headers3 = ['Vehicle', 'Price', 'Speed Limit', 'Maximum Group Size', 'Crew Size', 'Vehicle Size', 'Passenger Experience']
# for index, header in enumerate(headers3, 1):
#     vehicle_ws.cell(row=1, column=index, value=header)

# #-----------------OPENPYXL API CONNECTION & DATA PULLING-----------------


    
# #Starship Worksheet
# starships = []

# response2 = requests.get(starshipAPI)
# starshipData = response2.json()
# starships += starshipData['results']

# while starshipData['next']:
#     response2 = requests.get(starshipData['next'])
#     starshipData = response2.json()
#     starships += starshipData['results']
    

# #Vehicle Worksheet
# vehicles = []

# response3 = requests.get(vehicleAPI)
# vehicleData = response3.json()
# vehicles += vehicleData['results']


#CODE: 



# data = [                  #here, I want to pull from below API datasets. 
#    ["Bobby",26,20000],
#    ["Harrish",31,23000],
#    ["Jaya",18,19000],
#    ["Mark",22, 20500],
# ]
# index=0
# def read_data(): #reads the data and pulls from it
#    for index, line in enumerate(data):
#       tree.insert('', tk.END, iid = index,
#          text = line[0], values = line[1:])
# columns = ("Planet", "Climate", "Terrain", "Population", "Film Count")

# tree= ttk.Treeview(root, columns=columns ,height = 20)
# tree.pack(padx = 5, pady = 5)

# tree.heading('name', text='Planet') #list assigned to column 
# tree.heading('climate', text='Climate')
# tree.heading('terrain', text='Terrain')
# tree.heading('population', text='Population')
# tree.heading('film_count', text='Film Count')

# read_data()
# root.mainloop()



#COMBOBOX: used to select from a list of values. The values can be a list of strings or a list of numbers. The user can select a value from the drop-down list, which appears at the user's request.
    #from tkinter import ttk
        #combo = ttk.Combobox(master, values........)
    # top = Tk()
    # top.geometry("200x150")

    # frame = Frame(top)
    # frame.pack()

    # planet_elements = ["Climate", "Terrain", "Population", "Popularity"]

    # Combo = ttk.Combobox(frame, values = planet_elements)
    # Combo.set("What matters most to you when selecting a vacation spot?")
    # Combo.pack(padx = 5, pady = 5)
    # top.mainloop()


#-----------------OPENPYXL WORKBOOK SETUP-----------------+6

wb = Workbook()
planet_ws = wb.active
planet_ws.title = "Star Wars Vacation Spots"

starship_ws = wb.create_sheet("Getting There | Starships")

vehicle_ws = wb.create_sheet("Getting Around | Vehicles")

#-----------------OPENPYXL HEADER SETUP-----------------

#Planet Worksheet
headers = ['Planet', 'Climate', 'Terrain', 'Population', 'Film Count']
for index, header in enumerate(headers, 1):
    planet_ws.cell(row=1, column=index, value=header)

#Starship Worksheet
headers2 = ['Starship', 'Type of Ship', 'Price', 'Maximum Speed', 'Total Seats', 'Crew Size', 'Ship Size', 'Passenger Experience']
for index, header in enumerate(headers2, 1):
    starship_ws.cell(row=1, column=index, value=header)

#Vehicle Worksheet
headers3 = ['Vehicle', 'Price', 'Speed Limit', 'Maximum Group Size', 'Crew Size', 'Vehicle Size', 'Passenger Experience']
for index, header in enumerate(headers3, 1):
    vehicle_ws.cell(row=1, column=index, value=header)

#-----------------OPENPYXL API CONNECTION & DATA PULLING-----------------

#Planet Worksheet
planets = []

response = requests.get(planetAPI)
planetData = response.json()
planets += planetData['results']

while planetData['next']:
    response = requests.get(planetData['next'])
    planetData = response.json()
    planets += planetData['results']
    
#Starship Worksheet
starships = []

response2 = requests.get(starshipAPI)
starshipData = response2.json()
starships += starshipData['results']

while starshipData['next']:
    response2 = requests.get(starshipData['next'])
    starshipData = response2.json()
    starships += starshipData['results']
    
#Added to remove starships without passenger capacity                           #Mionne's code
starships[:] = (ship for ship in starships if ship['passengers'] != "0")        #Mionne's code
starships[:] = (ship for ship in starships if ship['passengers'] != "n/a")      #Mionne's code
starships[:] = (ship for ship in starships if ship['passengers'] != "unknown")  #Mionne's code

#Vehicle Worksheet
vehicles = []

response3 = requests.get(vehicleAPI)
vehicleData = response3.json()
vehicles += vehicleData['results']

while vehicleData['next']:
    response3 = requests.get(vehicleData['next'])
    vehicleData = response3.json()
    vehicles += vehicleData['results']
    
#Added to remove vehicles without passenger capacity                        #Mionne's code
vehicles[:] = (veh for veh in vehicles if veh['passengers'] != "0")         #Mionne's code
vehicles[:] = (veh for veh in vehicles if veh['passengers'] != "n/a")       #Mionne's code
vehicles[:] = (veh for veh in vehicles if veh['passengers'] != "unknown")   #Mionne's code

#-----------------OPENPYXL POPULATING DATA IN 'STAR WARS VACATION SPOTS' WORKSHEET-----------------

#Planet Worksheet
for row_index, planet in enumerate(planets, start=2):
    for col_index, header in enumerate(headers, start=1):
        if header == "Planet":
            value = planet.get("name")
        elif header == "Climate":
            value = planet.get("climate")
        elif header == "Terrain":
            value = planet.get("terrain")
        elif header == "Population":
            value = planet.get("population")
            if value == "unknown" or value == "":
                value = "Unknown"
        elif header == "Film Count":
            value = len(planet.get("films"))
        planet_ws.cell(row=row_index, column=col_index, value=value)

#Starship Worksheet
for row_index, starship in enumerate(starships, start=2):
    for col_index, header in enumerate(headers2, start=1):
        if header == "Starship":
            value = starship.get("name")
        elif header == "Type of Ship":
            value = starship.get("starship_class")
        elif header == "Price":
            value = starship.get("cost_in_credits")
        elif header == "Maximum Speed":
            value = starship.get("max_atmosphering_speed")
            if value == "Unknown" or value == [] or value == "" or value == "n/a":
                value = "This will be a slow ride. The starship ride IS your vacation."
        elif header == "Total Seats":
            value = starship.get("passengers")
        elif header == "Crew Size":
            value = starship.get("crew")
        elif header == "Ship Size":
            value = starship.get("length", "Unknown Size")
            try:
                value_float = float(value) #TRYing to convert value to a float (float because we may have decimals somewhere in the data)
                if 1 <= value_float <= 20:
                    value = "Small"
                elif 20.1 <= value_float <= 500:
                    value = "Medium"
                elif 500.1 <= value_float <= 4000:
                    value = "Large"
                elif 4000.1 <= value_float:
                    value = "Massive"
            except ValueError: #if value is not a number and pops an error, it will be a string that prints:
                value = "Unknown Size" #We can change this to anything we want it to say in cell on the worksheet.
        elif header == "Passenger Experience":          #NOTE: I used the cruise ship crew to passenger ratio as a guide: passengers / crew = how many passengers each crew member is responsible for. 1:1 is top notch. 
            passengers = starship.get("passengers", "0") #Defaults to 0 if no passengers
            crew = starship.get("crew", "1") #Defaults to 1 if no crew (this is IMPORTANT: you cannot divide by 0)
            if passengers.isnumeric() and crew.isnumeric(): #Checks if passengers and crew are numbers, does not convert anything
                passengers = int(passengers) #Converts passengers to an int
                crew = int(crew) #Converts crew to an int
                if crew > 0:
                    ratio = passengers / crew   #cruise ship quality of care ratio formula
                    if ratio <= 1.0:
                        value = "Luxury"
                    elif ratio >= 1.1 and ratio <= 2.9:
                        value = "Comfort"
                    elif ratio >= 3:
                        value = "Economy"
                else:
                    value = "Risky Experience"
            else:
                value = "Mystery Experience"    
        starship_ws.cell(row=row_index, column=col_index, value=value)

#Vehicle Worksheet
for row_index, vehicle in enumerate(vehicles, start=2):
    for col_index, header in enumerate(headers3, start=1):
        if header == "Vehicle":
            value = vehicle.get("model")
        elif header == "Price":
            value = vehicle.get("cost_in_credits")
        elif header == "Speed Limit":
            value = vehicle.get("max_atmosphering_speed")
        elif header == "Maximum Group Size":
            value = vehicle.get("passengers")
        elif header == "Crew Size":
            value = vehicle.get("crew")
        elif header == "Vehicle Size":
            value = vehicle.get("length", "Unknown Size")
            try:
                value_float = float(value) #TRYing to convert value to a float
                if 1 <= value_float <= 20:
                    value = "Small"
                elif 20.1 <= value_float <= 500:
                    value = "Medium"
                elif 500.1 <= value_float <= 4000:
                    value = "Large"
                elif 4000.1 <= value_float:
                    value = "Massive"
            except ValueError: #if value is not a number, it will be a string
                value = "Unknown Size" #we could also assign a default value here if the value cannot be converted to a float.
        elif header == "Passenger Experience":          #NOTE: I used cruise ship crew to passenger ratio as a guide: passengers / crew = how many passengers each crew member is responsible for. 1:1 is top notch.
            passengers = vehicle.get("passengers", "0") #Defaults to 0 if no passengers
            crew = vehicle.get("crew", "1") #Defaults to 1 if no crew data (this is important as you cannot divide by 0)
            if passengers.isnumeric() and crew.isnumeric(): #Checks if passengers and crew are numbers, does not convert anything
                passengers = int(passengers) #Converts passengers to an int
                crew = int(crew) #Converts crew to an int
                if crew > 0:
                    ratio = passengers / crew #cruise ship quality of care ratio formula
                    if ratio <= 1.0:
                        value = "Luxury"
                    elif ratio >= 1.1 and ratio <= 2.9:
                        value = "Comfort"
                    elif ratio >= 3:
                        value = "Economy"
                else:
                    value = "Risky Experience"
            else:
                value = "Mystery Experience"   

        vehicle_ws.cell(row=row_index, column=col_index, value=value)
                
#-----------------OPENPYXL SAVE-----------------

wb.save("spreadsheets/StarWarsVacationSpots.xlsx")

#-----------------GUI-----------------

# class App(tk.Tk):             #FIXME: probably remove
#     def __init__(self):       #FIXME: probably remove
#         super().__init__()    #FIXME: probably remove
root = tk.Tk()
nb = ttk.Notebook(root) #NOTEBOOK WIDGET

#app = App()                    #FIXME: probably remove
#app.mainloop()                 #FIXME: probably remove

#-----------------GUI|WIDGETS-----------------

#------------------<NOTEBOOK WIDGET>-----------------------
#notebook = ttk.Notebook(master, *options) <syntax for notebook widget>

#FRAMES : Planets, Starships, Vehicles (our tabs for our notebook)
planet_frame = ttk.Frame(nb)
starship_frame = ttk.Frame(nb)
vehicle_frame = ttk.Frame(nb)

#Planets
planet_label = ttk.Label(planet_frame, text="Choose a Planet")
planet_label.pack(pady = 50, padx = 20)

planet_headers = ('Planet', 'Climate', 'Terrain', 'Population', 'Film Count')
tree= ttk.Treeview(root, columns=planet_headers , height = 20)
tree.pack(padx = 5, pady = 5)

tree.add_packing_tree = ttk.Treeview(root, columns=planet_headers , height = 20)
tree.add_packing_tree['columns'] = ('Planet', 'Climate', 'Terrain', 'Population', 'Film Count')
tree.add_packing_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
tree.add_packing_tree.column('#0', anchor=tk.W, minwidth=0, stretch=0, width=0)
tree.add_packing_tree.column('Planet', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Climate', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Terrain', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Population', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Film Count', anchor=tk.W, minwidth=50, stretch=1, width=50)

tree.add_packing_tree.heading('Planet', text="Planet")
tree.add_packing_tree.heading('Climate', text="Climate")
tree.add_packing_tree.heading('Terrain', text="Terrain")
tree.add_packing_tree.heading('Population', text="Population")
tree.add_packing_tree.heading('Film Count', text="Film Count")
#Pulls Planet Data into above columns and rows within Treeview Widget
for row in planet_ws.iter_rows(min_row=2, min_col=0):
    for cell in row:
        tree.add_packing_tree.insert('', 'end', values=[cell.value for cell in row])


#Starships
starship_label = ttk.Label(starship_frame, text="Getting There | Starship")
starship_label.pack(pady = 50, padx = 20)

starship_headers = ('Starship', 'Type of Ship', 'Price', 'Maximum Speed', 'Total Seats', 'Crew Size', 'Ship Size', 'Passenger Experience')
tree= ttk.Treeview(root, columns=starship_headers , height = 20)
tree.pack(padx = 5, pady = 5)

tree.add_packing_tree = ttk.Treeview(root, columns=starship_headers , height = 20)
tree.add_packing_tree['columns'] = ('Starship', 'Type of Ship', 'Price', 'Maximum Speed', 'Total Seats', 'Crew Size', 'Ship Size', 'Passenger Experience')
tree.add_packing_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
tree.add_packing_tree.column('#0', anchor=tk.W, minwidth=0, stretch=0, width=0)
tree.add_packing_tree.column('Starship', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Type of Ship', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Price', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Maximum Speed', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Total Seats', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Crew Size', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Ship Size', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Passenger Experience', anchor=tk.W, minwidth=50, stretch=1, width=50)

tree.add_packing_tree.heading('Starship', text="Starship")
tree.add_packing_tree.heading('Type of Ship', text="Type of Ship")
tree.add_packing_tree.heading('Price', text="Price")
tree.add_packing_tree.heading('Maximum Speed', text="Maximum Speed")
tree.add_packing_tree.heading('Total Seats', text="Total Seats")
tree.add_packing_tree.heading('Crew Size', text="Crew Size")
tree.add_packing_tree.heading('Ship Size', text="Ship Size")
tree.add_packing_tree.heading('Passenger Experience', text="Passenger Experience")
#Pulls Starship Data into above columns and rows within Treeview Widget
for row in starship_ws.iter_rows(min_row=2, min_col=0):
    for cell in row:
        tree.add_packing_tree.insert('', 'end', values=[cell.value for cell in row])

#Vehicles
vehicle_label = ttk.Label(vehicle_frame, text="Getting Around | Vehicle")
vehicle_label.pack(pady = 50, padx = 20)

vehicle_headers = ('Vehicle', 'Price', 'Speed Limit', 'Maximum Group Size', 'Crew Size', 'Vehicle Size', 'Passenger Experience')
tree= ttk.Treeview(root, columns=vehicle_headers , height = 20)
tree.pack(padx = 5, pady = 5)

tree.add_packing_tree = ttk.Treeview(root, columns=vehicle_headers , height = 20)
tree.add_packing_tree['columns'] = ('Vehicle', 'Price', 'Speed Limit', 'Maximum Group Size', 'Crew Size', 'Vehicle Size', 'Passenger Experience')
tree.add_packing_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
tree.add_packing_tree.column('#0', anchor=tk.W, minwidth=0, stretch=0, width=0)
tree.add_packing_tree.column('Vehicle', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Price', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Speed Limit', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Maximum Group Size', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Crew Size', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Vehicle Size', anchor=tk.W, minwidth=50, stretch=1, width=50)
tree.add_packing_tree.column('Passenger Experience', anchor=tk.W, minwidth=50, stretch=1, width=50)

tree.add_packing_tree.heading('Vehicle', text="Vehicle")
tree.add_packing_tree.heading('Price', text="Price")
tree.add_packing_tree.heading('Speed Limit', text="Speed Limit")
tree.add_packing_tree.heading('Maximum Group Size', text="Maximum Group Size")
tree.add_packing_tree.heading('Crew Size', text="Crew Size")
tree.add_packing_tree.heading('Vehicle Size', text="Vehicle Size")
tree.add_packing_tree.heading('Passenger Experience', text="Passenger Experience")
#Pulls Vehicle Data into above columns and rows within Treeview Widget
for row in vehicle_ws.iter_rows(min_row=2, min_col=0):
    for cell in row:
        tree.add_packing_tree.insert('', 'end', values=[cell.value for cell in row])

#Pack Frames
planet_frame.pack(fill= tk.BOTH, expand=True)
starship_frame.pack(fill= tk.BOTH, expand=True)
vehicle_frame.pack(fill= tk.BOTH, expand=True)
#Add Frames to Notebook
nb.add(planet_frame, text="Planets")
nb.add(starship_frame, text="Starships")
nb.add(vehicle_frame, text="Vehicles")
#Pack Notebook
#root.mainloop() <moved to bottom of GUI code>

#------------------<RESIZE WINDOW WIDGET>-----------------------

root.geometry("500x700") #<syntax for resize window widget>
frame = ttk.Frame(root)
#label = ttk.Label(root, text = "So you want to travel the Star Wars universe and don't know where to begin?")
#label.pack(padx = 5, pady = 5)
sizegrip = ttk.Sizegrip(frame)
sizegrip.pack(expand = True, fill = tk.BOTH, anchor = tk.SE)
frame.pack(padx = 10, pady = 10, expand = True, fill = tk.BOTH)

#Pack Notebook
root.mainloop()
StopIteration